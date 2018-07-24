# -*- coding: utf-8 -*-
"""
Created on Tue Dec 26 17:26:07 2017

@author: Han
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 23:36:36 2017

"""

#!/usr/bin/env python
import urllib
import urllib.request
from bs4 import BeautifulSoup
import re,os
from openpyxl import Workbook,load_workbook

httpregex = re.compile('https?:[^\s]*')

def open_page(page,header,pagenumber):
    global ws,postid_list,postid_name,postlen

    req = urllib.request.Request(page, headers=header)
    html = urllib.request.urlopen(req)
    soup = BeautifulSoup(html.read(), "lxml")
    postlist = soup.find_all('article',id=re.compile('post'))
    for post in postlist:
        postid = int(post['id'][5:])
        if postid == 676:
            continue
        imgaelink = post.find('a',class_='post-cover')
        try:
           postlink = post.header.h2.a['href']
        except Exception:
           postlink = []

        if postlink:
            try:
                baidu_link = get_baidu_link(postlink)
                if baidu_link:
                    postname = baidu_link[-1]
                    if postid in postid_list:
                         index = postid_list.index(postid)
                         if postid_name[index] == postname:
                             print("%d exist "%(postid,))
                             if postid < 4100:
                                return False
                             else:
                                continue
                    #         continue
                    #         return False
                    
                    postlen = postlen+1
                    postid_list.append(postid)
                    postid_name.append(postname)
                    ws.cell(row=postlen,column=1,value=postid)
                    for ii in range(2,6):
                        ws.cell(row=postlen,column=ii,value=baidu_link[ii-2])
                #    if imgaelink:
                #        imgaelink = imgaelink.img['src']
                #        figreq = urllib.request.Request(imgaelink)
                #        fightml = urllib.request.urlopen(figreq).read()
                #        binfile = open('%s/%d.jpg' % ('Cover',postid ) , "wb")
                #        binfile.write(fightml);
                #        binfile.close();
    
                    try:
                        pass
                    except Exception :
                            print(imgaelink+" error")
                            pass
            except Exception :
                print(postlink+" error")
                pass



    print('Finish page: %d'%(pagenumber,))
    nextpage = soup(text='下一页')
    if nextpage:
        try:
           nextlink = nextpage[0].parent['href']
        except Exception:
           print('No Next Page on url:%s'%(page))
           return False
        pagenumber = pagenumber+1
        return (nextlink,pagenumber)
    else:
        return False


def get_baidu_link(page):
    #page = 'https://fulifb.com/2790.html'
    number = page.split('/')[-1]
    number = number[:-5]
    req = urllib.request.Request(page, headers=header)
    html = urllib.request.urlopen(req)
    soup = BeautifulSoup(html.read(), "lxml")
    name = soup.title.contents[0]
    a=soup.head.find('meta',content=re.compile("baidu"))
    if a:
        text = a['content']
        text = text[text.find('【'):]
        baidulink = httpregex.findall(text)
        if baidulink:
            baidulink = baidulink[0]
            text = text.replace(baidulink,'')
            text = text.replace('erphpdown','')
            passwd = re.findall('[a-zA-Z\d]{4}',text)
            a = (passwd[0],baidulink,passwd[1],name)
            print('%s  Album: %s'%(number,name,))
            return a
    return False

if ( __name__ == '__main__' ):
    cwd = os.getcwd()
    print(cwd)
    os.chdir('/root/tools/adb123')
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36"
        , "Connection": "keep-alive"
        }
    homepage = 'https://fulifb.com/'
    if not os.path.isdir('Cover'):
        os.mkdir('Cover')

    if os.path.isfile('list.xlsx'):
        wb = load_workbook('list.xlsx')
        ws = wb.active
        postid_list = [ i.value for i in ws['A']]
        postid_name = [ i.value for i in ws['E']]
        postlen = len(postid_list)
    else:
        wb = Workbook()
        ws = wb.active
        postid_list =[]
        postid_name = []
        postlen = 0


    page = homepage
    pagenumber = 1
    while True:
       a = open_page(page,header,pagenumber)
       if a:
           page = a[0]
           pagenumber = a[1]
       else:
           break
       wb.save('list.xlsx')
  
    index = [i[0] for i in sorted(enumerate(postid_list), key=lambda x:-x[1])] 

    wb2 = Workbook()
    ws2 = wb2.active

    for i,j in enumerate(index):
        oldrow = j+1
        newrow = i+1
        for k,data in enumerate(ws[oldrow]):
            d=ws2.cell(row=newrow, column=k+1,value=data.value)
        
    wb2.save('list.xlsx')   
    os.chdir(cwd)
